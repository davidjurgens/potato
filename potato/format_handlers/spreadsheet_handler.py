"""
Spreadsheet Format Handler

Extracts data from spreadsheet files (Excel, CSV, TSV) with row/cell
coordinate mapping for annotation.

Usage:
    from potato.format_handlers.spreadsheet_handler import SpreadsheetHandler

    handler = SpreadsheetHandler()
    output = handler.extract("data.xlsx", {
        "annotation_mode": "row",  # or "cell"
        "max_rows": 1000,
    })
"""

from typing import Dict, List, Any, Optional
from pathlib import Path
import html
import logging
import csv

from .base import BaseFormatHandler, FormatOutput
from .coordinate_mapping import (
    CoordinateMapper,
    SpreadsheetCoordinate,
    get_cell_reference,
)

logger = logging.getLogger(__name__)

# Check if dependencies are available
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    pd = None


class SpreadsheetHandler(BaseFormatHandler):
    """
    Handler for spreadsheet files.

    Supports Excel (.xlsx, .xls) via openpyxl and CSV/TSV via pandas or stdlib.
    Provides row-based and cell-based annotation modes.
    """

    format_name = "spreadsheet"
    supported_extensions = [".csv", ".tsv", ".xlsx", ".xls"]
    description = "Spreadsheet extraction with row/cell coordinate mapping"
    requires_dependencies = ["openpyxl"]

    def get_default_options(self) -> Dict[str, Any]:
        """Get default extraction options."""
        return {
            "annotation_mode": "row",  # "row", "cell", or "range"
            "max_rows": 1000,
            "header_row": 0,  # Row index for headers (None for no headers)
            "sheet_name": None,  # Sheet to extract (None for first/active)
            "skip_empty_rows": True,
            "text_columns": None,  # Columns to include (None for all)
            "row_separator": "\n",
            "cell_separator": "\t",
        }

    def extract(
        self,
        file_path: str,
        options: Optional[Dict[str, Any]] = None
    ) -> FormatOutput:
        """
        Extract data from a spreadsheet file.

        Args:
            file_path: Path to the spreadsheet file
            options: Extraction options:
                - annotation_mode: "row" (annotate rows) or "cell" (annotate cells)
                - max_rows: Maximum rows to process
                - header_row: Row index for column headers
                - sheet_name: Sheet to extract (for Excel files)

        Returns:
            FormatOutput with extracted text, HTML table, and coordinate mappings
        """
        opts = self.merge_options(options)
        path = Path(file_path)
        ext = path.suffix.lower()

        # Load data based on file type
        if ext in [".xlsx", ".xls"]:
            if not OPENPYXL_AVAILABLE:
                raise ImportError(
                    "openpyxl is required for Excel extraction. "
                    "Install with: pip install openpyxl"
                )
            data, headers, sheet_name = self._load_excel(file_path, opts)
        else:
            # CSV/TSV
            data, headers = self._load_csv(file_path, opts)
            sheet_name = None

        # Process data
        mapper = CoordinateMapper()
        text_parts = []
        html_parts = []
        current_offset = 0

        metadata = {
            "format": "spreadsheet",
            "source_file": str(file_path),
            "file_type": ext[1:],  # Remove dot
            "row_count": len(data),
            "column_count": len(headers) if headers else (len(data[0]) if data else 0),
            "headers": headers,
            "sheet_name": sheet_name,
            "annotation_mode": opts["annotation_mode"],
        }

        # Build HTML table
        html_parts.append(
            f'<div class="spreadsheet-container" data-mode="{opts["annotation_mode"]}">'
        )
        html_parts.append('<table class="spreadsheet-table">')

        # Header row
        if headers:
            html_parts.append('<thead><tr>')
            for col_idx, header in enumerate(headers):
                html_parts.append(f'<th data-col="{col_idx}">{html.escape(str(header))}</th>')
            html_parts.append('</tr></thead>')

        # Data rows
        html_parts.append('<tbody>')

        row_separator = opts["row_separator"]
        cell_separator = opts["cell_separator"]

        for row_idx, row in enumerate(data):
            row_start = current_offset
            row_texts = []

            html_parts.append(
                f'<tr class="spreadsheet-row" data-row="{row_idx}" '
                f'data-start="{row_start}">'
            )

            for col_idx, cell_value in enumerate(row):
                cell_text = str(cell_value) if cell_value is not None else ""
                cell_start = current_offset
                cell_end = cell_start + len(cell_text)

                row_texts.append(cell_text)

                # Add cell coordinate mapping
                cell_ref = get_cell_reference(row_idx, col_idx)

                if opts["annotation_mode"] == "cell":
                    mapper.add_mapping(
                        cell_start,
                        cell_end,
                        SpreadsheetCoordinate(
                            row=row_idx,
                            col=col_idx,
                            cell_ref=cell_ref,
                            sheet=sheet_name,
                        )
                    )

                # Build cell HTML
                data_attrs = (
                    f'data-row="{row_idx}" '
                    f'data-col="{col_idx}" '
                    f'data-cell-ref="{cell_ref}" '
                    f'data-start="{cell_start}" '
                    f'data-end="{cell_end}"'
                )
                html_parts.append(
                    f'<td class="spreadsheet-cell" {data_attrs}>'
                    f'{html.escape(cell_text)}</td>'
                )

                current_offset = cell_end
                if col_idx < len(row) - 1:
                    current_offset += len(cell_separator)

            html_parts.append('</tr>')

            # Build row text
            row_text = cell_separator.join(row_texts)
            text_parts.append(row_text)

            row_end = current_offset

            # Add row coordinate mapping
            if opts["annotation_mode"] == "row":
                mapper.add_mapping(
                    row_start,
                    row_end,
                    SpreadsheetCoordinate(
                        row=row_idx,
                        sheet=sheet_name,
                    )
                )

            # Add row separator
            current_offset += len(row_separator)

        html_parts.append('</tbody>')
        html_parts.append('</table>')
        html_parts.append('</div>')

        full_text = row_separator.join(text_parts)
        full_html = "\n".join(html_parts)

        coord_dict = mapper.to_dict()
        coord_dict["get_coords_for_range"] = mapper.get_coords_for_range

        return FormatOutput(
            text=full_text,
            rendered_html=full_html,
            coordinate_map=coord_dict,
            metadata=metadata,
            format_name=self.format_name,
            source_path=str(file_path),
        )

    def _load_excel(
        self,
        file_path: str,
        opts: Dict[str, Any]
    ) -> tuple:
        """
        Load data from an Excel file using openpyxl.

        Returns:
            Tuple of (data_rows, headers, sheet_name)
        """
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        # Select sheet
        sheet_name = opts.get("sheet_name")
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(
                    f"Sheet '{sheet_name}' not found. "
                    f"Available: {', '.join(wb.sheetnames)}"
                )
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title

        # Read data
        data = []
        headers = None
        header_row = opts.get("header_row")
        max_rows = opts.get("max_rows", 1000)
        skip_empty = opts.get("skip_empty_rows", True)
        text_columns = opts.get("text_columns")

        for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows + (1 if header_row is not None else 0))):
            row_values = [cell.value for cell in row]

            # Filter columns if specified
            if text_columns:
                row_values = [row_values[i] for i in text_columns if i < len(row_values)]

            # Check for empty rows
            if skip_empty and all(v is None or str(v).strip() == "" for v in row_values):
                continue

            # Handle header row
            if header_row is not None and row_idx == header_row:
                headers = [str(v) if v else f"Column_{i}" for i, v in enumerate(row_values)]
                continue

            data.append(row_values)

            if len(data) >= max_rows:
                break

        wb.close()
        return data, headers, sheet_name

    def _load_csv(
        self,
        file_path: str,
        opts: Dict[str, Any]
    ) -> tuple:
        """
        Load data from a CSV/TSV file.

        Returns:
            Tuple of (data_rows, headers)
        """
        path = Path(file_path)
        ext = path.suffix.lower()

        # Determine delimiter
        delimiter = "\t" if ext == ".tsv" else ","

        data = []
        headers = None
        header_row = opts.get("header_row")
        max_rows = opts.get("max_rows", 1000)
        skip_empty = opts.get("skip_empty_rows", True)
        text_columns = opts.get("text_columns")

        # Try pandas first if available
        if PANDAS_AVAILABLE:
            try:
                df = pd.read_csv(
                    file_path,
                    delimiter=delimiter,
                    header=header_row,
                    nrows=max_rows,
                    skip_blank_lines=skip_empty,
                    usecols=text_columns,
                )
                headers = df.columns.tolist()
                data = df.values.tolist()
                return data, headers
            except Exception as e:
                logger.debug(f"Pandas read failed, falling back to csv: {e}")

        # Fall back to stdlib csv
        with open(file_path, "r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)

            for row_idx, row in enumerate(reader):
                # Filter columns if specified
                if text_columns:
                    row = [row[i] for i in text_columns if i < len(row)]

                # Check for empty rows
                if skip_empty and all(v.strip() == "" for v in row):
                    continue

                # Handle header row
                if header_row is not None and row_idx == header_row:
                    headers = [v if v else f"Column_{i}" for i, v in enumerate(row)]
                    continue

                data.append(row)

                if len(data) >= max_rows:
                    break

        return data, headers

    def get_sheet_names(self, file_path: str) -> List[str]:
        """
        Get list of sheet names in an Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            List of sheet names
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required")

        wb = openpyxl.load_workbook(file_path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names

    def extract_sheet(
        self,
        file_path: str,
        sheet_name: str,
        options: Optional[Dict[str, Any]] = None
    ) -> FormatOutput:
        """
        Extract a specific sheet from an Excel file.

        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet to extract
            options: Extraction options

        Returns:
            FormatOutput for the specified sheet
        """
        opts = self.merge_options(options) if options else self.get_default_options()
        opts["sheet_name"] = sheet_name
        return self.extract(file_path, opts)
